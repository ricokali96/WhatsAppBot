import openpyxl
from IPython.display import display
import pandas as pd

#writer = pd.ExcelWriter('teste.xlsx', engine='openpyxl')
#wb =  writer.book

#wb.save('teste.xlsx')

#wb = load_workbook['teste.xlsx']
#ws = wb['Dadosauto']

#display(ws['A1'])

#writer = pd.ExcelWriter('teste.xlsx')
#for _,i in df.iterrows():
#    t_df = pd.DataFrame([i['Value']])
# ##   t_df.to_excel(writer, startcol=i['Column'],startrow=i['Row'], sheet_name=i['Sheet'], header=None, index=False)
#writer.save()

# esse funciona
sheetfile = openpyxl.load_workbook('teste.xlsx', read_only= False)
sheetname = sheetfile['Planilha1']
sheetname.cell(row=2,column=2).value = "ola tudo bem"

sheetfile.save('teste.xlsx')

#teste de novo algoritmo

"""
fn = 'teste.xlsx'

df = pd.read_excel(fn, header=None)
df2 = pd.DataFrame({'Data': [13, 24, 35, 46]})

writer = pd.ExcelWriter(fn, engine='openpyxl')
book = load_workbook(fn, read_only=False)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

#df.to_excel(writer, sheet_name='Sheet1', header=None, index=False)
df2.to_excel(writer, sheet_name='Sheet1', header=None, index=False,
             startcol=3,startrow=3)

writer.save()

"""