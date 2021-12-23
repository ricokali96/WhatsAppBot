from __future__ import print_function
from time import sleep, time
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from datetime import timedelta, date
from numpy import dtype, string_
import pandas as pd
from IPython.display import display
import openpyxl
import getpass
import random
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time, urllib, random, os
    ################
    #Funciona apenas para Windows
    ################
#Pega a data desejada
day_before = (date.today()-timedelta(days=2)).strftime("%d/%m/%Y")
day_beforestg= str(day_before)
print(day_before)

#cria o scopo - o que vai ser feito
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

#caminho do json com credenciais
SERVICE_ACCOUNT_FILE = 'keys.json'

creds = None
creds = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)

# The ID and range of a sample spreadsheet.
SAMPLE_SPREADSHEET_ID = '11tR34CLHoBnZLQQDMDunh0MJ3UsvyO_F4CMUxsmxI2Y'
SAMPLE_RANGE_NAME = 'Autocenter!A1000:M'


def main():
    """Shows basic usage of the Sheets API.
    Prints values from a sample spreadsheet.
    """
try:
    service = build('sheets', 'v4', credentials=creds)

    ## Call the Sheets API
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                                range=SAMPLE_RANGE_NAME).execute()
    #organiza os valores entre []                            
    values = result.get('values', [])
    #define nome das colunas de df1 e new_df
    colums_names = ['Mes','Data','Codigo','Vendedor','Nome','Condutor',
                                    'Telefone','Email','Km','Mecanico','Sexo','Marca','Modelo']
    df1 = pd.DataFrame(values, columns= colums_names) #cria dataframe com dados da api
    #inicializa a nova dataframe
    new_df = pd.DataFrame(data=None, index=None, columns=colums_names, dtype=None, copy=False)
    #display(df1.loc[1])
    ##Cria nova dataframe com linhas de data de hoje - timedelta
    j = 0
    for i in range(0, len(values)):
        if str(df1.iloc[i][1]) == day_beforestg:
            new_df.loc[j] = df1.loc[i]
            j = j + 1
    display(new_df)

    ##Rotina para exportar o dataframe com apenas o dia definido, para excel em linhas e colunas definidas
    sheetfile = openpyxl.load_workbook('TesteReal1.xlsx', read_only= False)
    sheetname = sheetfile['Planilha1'] #Nome da pagina da planilha
    for i, condutor in enumerate(new_df['Condutor']): #Loop para colocar na planilha as colunas definidas.
        sheetname.cell(row=(i+2),column=1).value = new_df.loc[i,"Data"]
        sheetname.cell(row=(i+2),column=2).value = condutor
        sheetname.cell(row=(i+2),column=3).value = new_df.loc[i,"Marca"]
        sheetname.cell(row=(i+2),column=4).value = new_df.loc[i,"Modelo"]
        sheetname.cell(row=(i+2),column=5).value = new_df.loc[i,"Telefone"]
    sheetfile.save('TesteReal1.xlsx')

    ##Abrir o excel para que as formulas das colunas sejam lidas ##So funciona no windows 
    ##//Para MAC e Linux ->>>> os.system('open TesteReal1.xlsx')
    #subprocess.call('C://Program Files//Microsoft Office//root//Office16//EXCEL.EXE')
    os.startfile('TesteReal1.xlsx')
    time.sleep(10)

    ##rotina para mandar as mensagens com chrome
    usern = getpass.getuser()
    profiledir = r"--user-data-dir=C:\Users\{}\AppData\Local\Google\Chrome\User Data\Default".format(usern)
    contatos_df = pd.read_excel("TesteReal1.xlsx")
    display(contatos_df)
    options = webdriver.ChromeOptions()
    #options.add_argument(r"--user-data-dir=C:\Users\Usuário\AppData\Local\Google\Chrome\User Data\Default")
    options.add_argument(profiledir)
    #while len(navegador.find_element_by_id("side")) < 1:
    #time.sleep(5)
    navegador = webdriver.Chrome(executable_path=r'./chromedriver', chrome_options=options)
    navegador.get("https://web.whatsapp.com/")
    time.sleep(random.randint(35,40)) #tempo para escanear QR code

    for i, mensagem in enumerate(contatos_df['Mensagem']):
        #dia = contatos_df.loc[i,"Data"]
        #if dia == date_time:
        pessoa = contatos_df.loc[i,"Nome"]
        numero = contatos_df.loc[i,"Numero"]
        #print(numero,"",pessoa,"",dia)
        textcod = urllib.parse.quote(mensagem) #transforma a mensagem em cod url, pode customisar com variaveis
        link = f"https://web.whatsapp.com/send?phone={numero}&text={textcod}"
        navegador.get(link)
        time.sleep(random.randint(20,25))
        #while len(navegador.find_element_by_id("side")) < 1:
        #time.sleep(1)
        try: #tenta apertar enter, se nao der avisa que deu errado
            navegador.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[2]').send_keys(Keys.ENTER) # copiamos o XPATH da caixa de texto e vamos dar enter
        except:
            print(f"Numero errado de {pessoa}, ver se {numero} está correto")
        time.sleep(random.randint(6,10)) #tempo para envio da mensagem

except HttpError as err:
    print(err)

if __name__ == '__main__':
    main()